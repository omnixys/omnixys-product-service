"""
Produkt-Lesedienst für den Product-Microservice mit optionaler Excel-Exportfunktion.

Diese Klasse kapselt alle lesenden Zugriffe auf MongoDB und bietet Funktionen zur Abfrage von Produkten
sowie zur Erstellung eines Excelsheets (falls aktiviert).
"""

import csv
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Final, List, Optional

from beanie import PydanticObjectId
from loguru import logger
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from opentelemetry import trace

from product.config import env
from product.config.feature_flags import excel_export_enabled  # z. B. True/False-Flag
from product.config.kafka import get_kafka_settings
from product.error.exceptions import NotFoundError
from product.logging.logger_plus import LoggerPlus
from product.messaging.kafka_singleton import get_kafka_producer
from product.messaging.producer import KafkaProducerService
from product.model.entity.product import Product, map_product_to_product_type
from product.repository.pageable import Pageable
from product.repository.product_repository import ProductRepository
from product.repository.slice import Slice
from product.tracing.trace_context import TraceContext
from product.tracing.trace_context_util import TraceContextUtil

tracer = trace.get_tracer(__name__)
as_csv: Final = env.EXPORT_FORMAT.lower() == "csv"


class ProductReadService:
    """Serviceklasse für lesenden Zugriff auf Produktdaten in MongoDB."""

    def __init__(self, repository: ProductRepository):
        self._repository = repository
        self._log = LoggerPlus()
        self._producer = get_kafka_producer()
        self._service = get_kafka_settings().client_id

    async def find_by_id(self, product_id: PydanticObjectId) -> Product:
        with tracer.start_as_current_span("ProductReadService.find_by_id"):
            await self._log.debug("find_by_id: id=%s", product_id)
            product = await self._repository.find_by_id_or_throw(product_id)
            await self.notify_export_event([product])
            return product

    async def notify_export_event(
        self, products: List[Product], action: str = "product_export"
    ) -> None:
        """
        Sendet ein Event an Kafka, das den erfolgreichen Produkt-Export signalisiert.
        """

        payload = {
            "event": action,
            "total_products": len(products),
            "exported_at": datetime.utcnow().isoformat(),
        }

        await self._producer.publish(
            topic="orders.cancelled",
            payload=payload,
            headers=[
                ("x-event-name", action),
                ("x-event-version", "1.0.0"),
                ("x-service", self._service),
            ],
        )

        await self._log.info("🛰️ Kafka-Export-Event versendet: %s", payload)

    async def find_all(self, pageable: Pageable) -> Slice:
        logger.debug("find_all")

        products = await self._repository.find_paginated(
            skip=pageable.skip,
            limit=pageable.limit,
        )

        if excel_export_enabled:
            ProductReadService._create_export_file(products)

        mapped = [map_product_to_product_type(p) for p in products]
        return Slice(
            content=mapped,
            total=len(mapped),
            size=pageable.limit,
            page=pageable.skip,
        )

    async def find_paginated(self, skip: int = 0, limit: int = 10) -> List[Product]:
        logger.debug("find_paginated: skip=%s, limit=%s", skip, limit)
        products = await self._repository.find_paginated(skip=skip, limit=limit)

        if excel_export_enabled:
            ProductReadService._create_export_file(products)
        return products

    async def find_filtered(self, filter_dict: dict, pageable: Pageable) -> Slice:
        logger.debug("find_filtered: filter_dict=%s", filter_dict)
        result = await self._repository.find_filtered(
            filter_dict,
            skip=pageable.skip,
            limit=pageable.limit,
        )

        if not result:
            logger.warning("Keine Produkte gefunden mit Filter: %s", filter_dict)
            raise NotFoundError("Keine Produkte mit diesen Filterkriterien gefunden.")

        if excel_export_enabled:
            ProductReadService._create_export_file(products=result)

        mapped = [map_product_to_product_type(p) for p in result]
        return Slice(
            content=mapped,
            total=len(mapped),
            pageable=pageable,
        )

    def _create_export_file(products: List[Product]) -> None:
        """Erstellt CSV oder Excel mit Logo und Diagrammen."""

        # 🔢 Startposition
        start_row = 10
        start_col = 2  # B = 2

        # 🔧 Konfiguration
        as_csv: Final = env.EXPORT_FORMAT.lower() == "csv"
        logger.debug("_create_export_file: as csv={}", as_csv)

        # ⏱ Zeitstempel für Dateinamen
        timestamp = datetime.now().strftime("%Y-%m-%d")
        path = Path("exports")
        path.mkdir(parents=True, exist_ok=True)
        file_name = f"{timestamp}.{'csv' if as_csv else 'xlsx'}"
        export_path = path / file_name

        # 🔄 Gemeinsame Zeilenstruktur
        header = [
            "Produkt-ID",
            "Name",
            "Marke",
            "Kategorie",
            "Preis (€)",
            "Tags",
            "Erstellt",
            "Geändert",
        ]
        rows = [
            [
                str(p.id),
                p.name,
                p.brand,
                p.category.value,
                float(p.price),
                ", ".join(p.tags or []),
                p.created.isoformat(),
                p.updated.isoformat(),
            ]
            for p in products
            if p.price > 0
        ]

        # ➤ CSV
        if as_csv:
            with open(export_path, mode="w", newline="", encoding="utf-8") as csv_file:
                writer = csv.writer(csv_file, delimiter=";")
                writer.writerow(header)
                writer.writerows(rows)
            logger.success("CSV-Export gespeichert unter: %s", export_path)
            return

        # 📊 Excel-Erstellung mit openpyxl
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Produkte"

        # Rahmen-Stile
        thick = Side(border_style="thick", color="000000")
        thin = Side(border_style="thin", color="000000")

        def table_border(row_idx: int, col_idx: int) -> Border:
            """Rahmenregel: außen dick, innen dünn."""

            is_top = row_idx == start_row
            is_bottom = row_idx in {start_row, start_row + len(rows)}
            is_left = col_idx == start_col
            is_right = col_idx == start_col + len(header) - 1

            return Border(
                top=thick if is_top else thin,
                bottom=thick if is_bottom else thin,
                left=thick if is_left else thin,
                right=thick if is_right else thin,
            )

        # Kopfzeile in Zeile 10 (B10)
        for col_offset, title in enumerate(header):
            col_idx = start_col + col_offset
            cell = sheet.cell(row=start_row, column=col_idx, value=title)
            cell.font = Font(bold=True)
            cell.border = table_border(start_row, col_idx)

        # 🔴 Zeilen mit Preis > 100 rot markieren
        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        for row_offset, row_data in enumerate(rows, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                row_idx = start_row + row_offset
                col_idx = col_index + (start_col - 1)
                cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = table_border(row_idx, col_idx)

                if col_index == 5 and cell_value > 100:
                    cell.fill = red_fill

        # 🖼️ Branding: Logo in Zelle A1
        # ➤ Logo einfügen (falls vorhanden)
        try:
            logo_path = Path(__file__).parent.parent / "static/logo.png"
            if logo_path.exists():
                logo = ExcelImage(str(logo_path))
                logo.width = 300
                logo.height = 80

                # 🔲 Zellen A1 bis C4 zusammenführen
                sheet.merge_cells("A1:C4")  # Logo-Bereich zusammenführen

                # 📏 Spaltenbreite und Zeilenhöhe für gute Darstellung
                sheet.column_dimensions["A"].width = 20
                sheet.column_dimensions["B"].width = 20
                sheet.column_dimensions["C"].width = 20
                sheet.row_dimensions[1].height = 40
                sheet.row_dimensions[2].height = 40
                sheet.row_dimensions[3].height = 40

                # ➕ Bild in A1 einfügen
                sheet.add_image(logo, "A1")
                print("🔍 Logo hinzugefügt von A1 bis C4")
            else:
                logger.warning("⚠️ Kein Logo gefunden: {}", logo_path)
        except Exception as e:
            logger.warning("⚠️ Fehler beim Logo: {}", str(e))

        # Anzahl Produkte pro Kategorie
        category_count = defaultdict(int)
        category_sum = defaultdict(float)
        category_prices = defaultdict(list)

        for p in products:
            key = p.category.value
            category_count[key] += 1
            category_sum[key] += float(p.price)
            category_prices[key].append((p.name, float(p.price)))

        # Kreisdiagramm: Anzahl Produkte pro Kategorie
        sheet_count = workbook.create_sheet("Anzahl je Kategorie")
        sheet_count.append(["Kategorie", "Anzahl"])
        for cat, count in category_count.items():
            sheet_count.append([cat, count])

        pie = PieChart()
        pie.title = "Produkte pro Kategorie"
        data = Reference(
            sheet_count, min_col=2, min_row=1, max_row=len(category_count) + 1
        )
        labels = Reference(
            sheet_count, min_col=1, min_row=2, max_row=len(category_count) + 1
        )
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        sheet_count.add_chart(pie, "E2")

        # Balkendiagramm: Preise pro Kategorie (Summe)
        sheet_sum = workbook.create_sheet("Preise je Kategorie")
        sheet_sum.append(["Kategorie", "Summe (€)"])
        for cat, total in category_sum.items():
            sheet_sum.append([cat, total])

        bar = BarChart()
        bar.title = "Summe der Preise je Kategorie"
        bar.x_axis.title = "Kategorie"
        bar.y_axis.title = "€"
        data = Reference(sheet_sum, min_col=2, min_row=1, max_row=len(category_sum) + 1)
        cats = Reference(sheet_sum, min_col=1, min_row=2, max_row=len(category_sum) + 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        sheet_sum.add_chart(bar, "E2")

        # Einzeldiagramme je Kategorie
        for cat, entries in category_prices.items():
            # ws = workbook.create_sheet(f"Preise: {cat[:20]}")
            ws = workbook.create_sheet(_sanitize_sheet_name(f"Preise {cat}"))
            ws.append(["Produktname", "Preis (€)"])
            for name, price in entries:
                ws.append([name, price])

            chart = BarChart()
            chart.title = f"Preise in Kategorie: {cat}"
            chart.x_axis.title = "Produkt"
            chart.y_axis.title = "Preis (€)"
            data = Reference(ws, min_col=2, min_row=1, max_row=len(entries) + 1)
            names = Reference(ws, min_col=1, min_row=2, max_row=len(entries) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(names)
            ws.add_chart(chart, "D2")

        # 💾 Speichern
        workbook.save(export_path)
        logger.success("📊 Excel gespeichert: {}", export_path)


def _sanitize_sheet_name(name: str) -> str:
    # Ungültige Excel-Zeichen entfernen oder ersetzen
    invalid_chars = r"[:\\/*?[\]]"
    name = re.sub(invalid_chars, "", name)
    return name[:31]  # Excel-Sheet-Titel dürfen max. 31 Zeichen haben
